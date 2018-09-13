from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string #列名数字化

wb = load_workbook("chemy.xlsx")#读取表格
#print(wb.sheetnames) 读取页面名

sheet = wb.get_sheet_by_name("Sheet1")#读取第一页

for rowOfCellObjects in sheet['C3':'BV74']:#定义遍历区间
    for i in rowOfCellObjects:
        if i.value == '×':
            a = i.row
            b = column_index_from_string(i.column)#列名数字化
            #print(a, b)
            #print(i.value, i.coordinate, i.row,i.column, sheet.cell(row=i.row, column=2).value)
            print('rule "1"')
            print('  salience 1')
            print('  When')
            print('$T:Type(type1=="'+sheet.cell(row=a, column=2).value+'",type2=="'+sheet.cell(row=b-1, column=2).value+'")')
            print('  Then')
            print('nums.add(number.builder().a("'+sheet.cell(row=a, column=2).value+'").b("'+sheet.cell(row=b-1, column=2).value+'").c(1).build());')
        else:
            print('rule "2"')
            print('  salience 2')
            print('  When')
            print('$T:Type(type1=="' + sheet.cell(row=i.row, column=2).value + '",type2=="' + sheet.cell(column_index_from_string(i.column) - 1, column=2).value + '")')
            print('  Then')
            print('nums.add(number.builder().a("' + sheet.cell(row=i.row, column=2).value + '").b("' + sheet.cell(column_index_from_string(i.column) - 1, column=2).value + '").c(0).build());')
            #print(sheet.cell(row=i.row, column=2).value + ' ' + sheet.cell(column_index_from_string(i.column) - 1, column=2).value + '0')
print('--- END OF ROW ---')
